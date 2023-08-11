from openpyxl import workbook, load_workbook


class utils():
    def __init__(self,driver):
        self.driver = driver

    def takeScreenshot(self, fileName):
        fileName = fileName+".png"
        self.driver.save_screenshot("Screenshot"+fileName)

    def readDataFromExcel(file_name, sheet):
        dataList = []
        wb = load_workbook(file_name)
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            dataList.append(row)
        return dataList


